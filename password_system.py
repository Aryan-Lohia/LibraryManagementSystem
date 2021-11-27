from openpyxl import Workbook,load_workbook
class password_management:
    data=load_workbook("PasswordDatabase.xlsx")
    database=data.active
    def __init__(self,username):
        self.username=username
    def new_user(self):
        password=input("Please set new password\n").lstrip().rstrip()
        self.database.append(list((self.username+" "+password).split()))
        print(f"Your password has been set.\nYour username is {self.username}.\nYour password is {password}\nPlease login with new Password.")
    def old_user(self,user_cell):
        password = input("Please enter password\n").lstrip().rstrip()
        if password == self.database["B"+str(user_cell+2)].value:
            print("Successful Login")
            return True
        else:
            print("Incorrect Password")
            print("Please Retry Login")
            return False

def password_sys(username):
    user = password_management(username)
    users_list=[]
    for rows in range(2,user.database.max_row+1):
        users_list.append(user.database["A" + str(rows)].value)
    if username in users_list:
        auth = user.old_user(users_list.index(username))
    else:
        user.new_user()
        auth = False
    user.data.save("PasswordDatabase.xlsx")
    return auth
